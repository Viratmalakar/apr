from flask import Flask, render_template, request, send_file, redirect, url_for
import pandas as pd
import os
import tempfile
from openpyxl import load_workbook

app = Flask(__name__)

# -----------------------------
# UNMERGE FUNCTION
# -----------------------------
def unmerge_excel(file):
    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb = load_workbook(file)

    for sheet in wb.worksheets:
        merged = list(sheet.merged_cells.ranges)
        for m in merged:
            sheet.unmerge_cells(str(m))

    wb.save(temp.name)
    return temp.name


# -----------------------------
# TIME FORMAT
# -----------------------------
def format_time(seconds):

    if pd.isna(seconds):
        return "00:00:00"

    seconds = int(seconds)
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60

    return f"{h:02}:{m:02}:{s:02}"


# -----------------------------
# GENERATE REPORT
# -----------------------------
@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/generate", methods=["POST"])
def generate():

    ap_file = request.files["agent_file"]
    cdr_file = request.files["cdr_file"]

    ap_path = unmerge_excel(ap_file)
    cdr_path = unmerge_excel(cdr_file)

    # -----------------------------
    # READ AGENT PERFORMANCE
    # -----------------------------
    ap = pd.read_excel(ap_path, skiprows=2)

    ap.replace("-", 0, inplace=True)

    ap["Employee ID"] = ap.iloc[:,1].astype(str)
    ap["Agent Name"] = ap.iloc[:,2]

    ap["Total Login"] = ap.iloc[:,3]

    ap["Total Break"] = (
        pd.to_numeric(ap.iloc[:,19], errors="coerce").fillna(0)
        +
        pd.to_numeric(ap.iloc[:,22], errors="coerce").fillna(0)
        +
        pd.to_numeric(ap.iloc[:,24], errors="coerce").fillna(0)
    )

    ap["Total Meeting"] = (
        pd.to_numeric(ap.iloc[:,20], errors="coerce").fillna(0)
        +
        pd.to_numeric(ap.iloc[:,23], errors="coerce").fillna(0)
    )

    ap["Total Net Login"] = ap["Total Login"] - ap["Total Break"]

    ap["Total Talk Time"] = pd.to_numeric(ap.iloc[:,5], errors="coerce").fillna(0)

    ap_final = ap[
        [
            "Employee ID",
            "Agent Name",
            "Total Login",
            "Total Net Login",
            "Total Break",
            "Total Meeting",
            "Total Talk Time"
        ]
    ]

    # -----------------------------
    # READ CDR
    # -----------------------------
    cdr = pd.read_excel(cdr_path, skiprows=1)

    cdr["Employee ID"] = cdr.iloc[:,1].astype(str)
    cdr["Campaign"] = cdr.iloc[:,6]
    cdr["Status"] = cdr.iloc[:,25]

    cdr["MatureFlag"] = cdr["Status"].isin(["CALLMATURED","TRANSFER"])
    cdr["IBFlag"] = (
        cdr["Campaign"].str.upper() == "CSRINBOUND"
    ) & cdr["MatureFlag"]

    total = cdr.groupby("Employee ID")["MatureFlag"].sum()
    ib = cdr.groupby("Employee ID")["IBFlag"].sum()

    mature = pd.DataFrame({

        "Employee ID": total.index,
        "Total Mature": total.values,
        "IB Mature": ib.values

    })

    mature["OB Mature"] = mature["Total Mature"] - mature["IB Mature"]

    # -----------------------------
    # MERGE
    # -----------------------------
    final = pd.merge(ap_final, mature, on="Employee ID", how="left")

    final.fillna(0, inplace=True)

    # -----------------------------
    # AHT
    # -----------------------------
    final["AHT"] = final.apply(

        lambda x: format_time(
            x["Total Talk Time"] / x["Total Mature"]
        ) if x["Total Mature"] > 0 else "00:00:00",

        axis=1

    )

    final["Total Login"] = final["Total Login"].apply(format_time)
    final["Total Net Login"] = final["Total Net Login"].apply(format_time)
    final["Total Break"] = final["Total Break"].apply(format_time)
    final["Total Meeting"] = final["Total Meeting"].apply(format_time)

    final = final[

        [
            "Employee ID",
            "Agent Name",
            "Total Login",
            "Total Net Login",
            "Total Break",
            "Total Meeting",
            "AHT",
            "Total Mature",
            "IB Mature",
            "OB Mature"
        ]

    ]

    final.to_excel("final_report.xlsx", index=False)

    return render_template(
        "result.html",
        tables=final.to_dict("records"),
        headers=final.columns
    )


# -----------------------------
# DOWNLOAD EXCEL
# -----------------------------
@app.route("/download")
def download():

    return send_file(

        "final_report.xlsx",
        as_attachment=True

    )


# -----------------------------
# RESET
# -----------------------------
@app.route("/reset")
def reset():
    return redirect("/")


# -----------------------------
# RUN
# -----------------------------
if __name__ == "__main__":

    app.run()
